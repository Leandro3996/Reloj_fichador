from rangefilter.filters import DateRangeFilter
from django.contrib import admin
from django.shortcuts import render, get_object_or_404, redirect
from django.db import models
from django.contrib.admin.widgets import AdminSplitDateTime
from django.contrib.admin.models import LogEntry
from django.core.management.base import BaseCommand
from django.utils import timezone
from django.utils.html import format_html
from .models import (
    Operario, RegistroDiario, Horas_trabajadas, Horas_extras, 
    Horas_totales, Area, Horario, Licencia, RegistroAsistencia, 
    Horas_feriado, HistoricalOperario, HistoricalRegistroDiario
)
from django.urls import path, reverse
from datetime import timedelta
from django.utils.translation import gettext_lazy as _
from .forms import LicenciaForm
from .utils import generar_pdf, generar_excel
from import_export.admin import ExportMixin
from datetime import datetime
from django.http import HttpResponse
import os
from django.conf import settings
from simple_history.admin import SimpleHistoryAdmin
from django.contrib.auth.models import User
from django.contrib.auth.admin import UserAdmin
from django.http import HttpResponseRedirect


class Command(BaseCommand):
    help = 'Genera registros de asistencia para los operarios activos'

    def handle(self, *args, **kwargs):
        hoy = timezone.now().date()
        operarios_activos = Operario.objects.filter(activo=True)

        for operario in operarios_activos:
            registro, created = RegistroAsistencia.objects.get_or_create(
                operario=operario,
                fecha=hoy
            )
            registro.verificar_asistencia()
            self.stdout.write(self.style.SUCCESS(f'Registro generado para {operario}'))

class ActivoInactivoFilter(admin.SimpleListFilter):
    title = _('Activo/Inactivo')
    parameter_name = 'activo'

    def lookups(self, request, model_admin):
        return (
            ('1', _('Activo')),
            ('0', _('Inactivo')),
        )

    def queryset(self, request, queryset):
        if self.value() == '1':
            return queryset.filter(activo=True)
        elif self.value() == '0':
            return queryset.filter(activo=False)
        return queryset

class LicenciaInline(admin.TabularInline):
    model = Licencia
    extra = 1
    fields = ['archivo', 'fecha_subida']
    readonly_fields = ['fecha_subida']

def exportar_pdf(modeladmin, request, queryset, calculate_hours_total=None):
    """
    Función genérica para exportar cualquier modelo a PDF.
    Determina automáticamente los campos y encabezados basados en el ModelAdmin
    
    Args:
        calculate_hours_total: Si es None, autodetecta. Si es True, fuerza el cálculo de totales.
                           Si es False, no calcula totales.
    """
    # Obtener lista de campos a mostrar del list_display del ModelAdmin
    if hasattr(modeladmin, 'list_display'):
        campos = [field for field in modeladmin.list_display]
    else:
        # Si no tiene list_display, usamos los campos definidos en el modelo
        campos = [field.name for field in modeladmin.model._meta.fields]
    
    # Generar los encabezados a partir de los nombres cortos de los campos o del HEADER_MAP si existe
    encabezados = []
    for campo in campos:
        if hasattr(modeladmin, 'HEADER_MAP') and campo in modeladmin.HEADER_MAP:
            encabezados.append(modeladmin.HEADER_MAP[campo])
        else:
            # Intentar obtener el verbose_name o short_description
            try:
                # Si es un método con short_description
                if hasattr(getattr(modeladmin, campo), 'short_description'):
                    encabezados.append(getattr(modeladmin, campo).short_description)
                # Si es un campo del modelo
                elif campo in [field.name for field in modeladmin.model._meta.fields]:
                    encabezados.append(modeladmin.model._meta.get_field(campo).verbose_name)
                else:
                    # Usar el nombre del campo capitalizado como fallback
                    encabezados.append(campo.replace('_', ' ').capitalize())
            except:
                # Si todo falla, usar el nombre del campo
                encabezados.append(campo.replace('_', ' ').capitalize())
    
    # Obtener el título del reporte basado en el verbose_name_plural del modelo
    titulo = f"Reporte de {modeladmin.model._meta.verbose_name_plural.capitalize()}"
    
    # Generar el PDF usando la función generar_pdf
    return generar_pdf(modeladmin, request, queryset, campos, encabezados, titulo, calculate_hours_total=calculate_hours_total)

exportar_pdf.short_description = "Exportar seleccionados a PDF"

# Mixin para añadir la acción de exportar a PDF
class ExportarPDFMixin:
    actions = ['exportar_pdf']
    
    def exportar_pdf(self, request, queryset):
        return exportar_pdf(self, request, queryset)
    exportar_pdf.short_description = "Exportar seleccionados a PDF"

@admin.register(Operario)
class OperarioAdmin(ExportMixin, SimpleHistoryAdmin):
    inlines = [LicenciaInline]
    list_display = (
        'dni', 'nombre', 'apellido', 'fecha_nacimiento', 'fecha_ingreso_empresa', 'titulo_tecnico', 'get_areas', 'activo', 'view_history_button'
    )
    list_filter = ('areas', ('fecha_nacimiento', DateRangeFilter), ('fecha_ingreso_empresa', DateRangeFilter), 'titulo_tecnico', ActivoInactivoFilter)
    search_fields = ('dni', 'nombre', 'apellido', 'fecha_nacimiento', 'fecha_ingreso_empresa', 'titulo_tecnico')
    filter_horizontal = ('areas',)
    actions = ['asignar_area', 'generar_reporte', 'exportar_excel', 'exportar_pdf']

    readonly_fields = ('foto_tag', 'get_areas')

    def get_areas(self, obj):
        return ", ".join([area.nombre for area in obj.areas.all()])
    get_areas.short_description = 'Áreas'

    def foto_tag(self, obj):
        if obj.foto:
            return format_html(f'<img src="{obj.foto.url}" style="max-width: 150px; height: auto;" />')
        return "(Sin foto)"
    foto_tag.short_description = 'Foto del operario'

    def view_history_button(self, obj):
        """Mostrar un botón para ver el historial del operario"""
        if obj.pk:
            url = reverse('admin:reloj_fichador_historicaloperario_changelist') + f"?id={obj.pk}"
            return format_html(
                '<a class="button" href="{}">Ver Historial</a>',
                url
            )
        return ""
    view_history_button.short_description = "Historial"

    def generar_reporte(self, request, queryset):
        registros = list(queryset)
        total_registros = Operario.objects.count()
        items_por_pagina = 30

        encabezados = ['DNI', 'Nombre', 'Apellido', 'Fecha Nacimiento', 'Fecha Ingreso',]

        filas = []
        for registro in registros:
            fila = [
                registro.dni,
                registro.nombre,
                registro.apellido,
                registro.fecha_nacimiento.strftime('%d/%m/%Y'),
                registro.fecha_ingreso_empresa.strftime('%d/%m/%Y'),               
            ]
            filas.append(fila)

        page_filas = [filas[i:i + items_por_pagina] for i in range(0, len(filas), items_por_pagina)]
        total_pages = len(page_filas)

        # Definir el título dinámico
        titulo_reporte = "Reporte de Operarios"

        context = {
            'encabezados': encabezados,
            'page_filas': page_filas,
            'num_columnas': len(encabezados),
            'current_date': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'current_page': 1,
            'total_pages': total_pages,
            'num_items_filtrados': len(registros),
            'total_items': total_registros,
            'titulo_reporte': titulo_reporte

        }

        return render(request, 'reloj_fichador/reporte.html', context)

    def exportar_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse

        registros = list(queryset)
        encabezados = ['DNI', 'Nombre', 'Apellido', 'Fecha Nacimiento', 'Fecha Ingreso',]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Operarios"
        sheet.append(encabezados)

        for registro in registros:
            fila = [
                registro.dni,
                registro.nombre,
                registro.apellido,
                registro.fecha_nacimiento.strftime('%d/%m/%Y'),
                registro.fecha_ingreso_empresa.strftime('%d/%m/%Y'),
            ]
            sheet.append(fila)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_de_Operarios.xlsx"'
        workbook.save(response)

        return response

    exportar_excel.short_description = "Exportar a Excel"


@admin.register(RegistroDiario)
class RegistroDiarioAdmin(ExportMixin, SimpleHistoryAdmin):
    list_display = ('get_dni', 'get_nombre', 'get_apellido', 'tipo_movimiento', 'formatted_hora_fichada', 
                    'origen_fichada', 'mostrar_inconsistencia', 'mostrar_valido', 'view_history_button')
    list_filter = ('inconsistencia','valido','tipo_movimiento', ('hora_fichada', DateRangeFilter),'origen_fichada',)
    search_fields = ('operario__dni', 'operario__nombre', 'operario__apellido')
    fields = ('operario', 'tipo_movimiento', 'hora_fichada', 'valido','descripcion_inconsistencia',)
    actions = ['generar_reporte', 'exportar_excel', 'exportar_pdf']      

    HEADER_MAP = {
        'get_dni': 'DNI',
        'get_nombre': 'Nombre',
        'get_apellido': 'Apellido',
        'tipo_movimiento': 'Tipo Movimiento',
        'formatted_hora_fichada': 'Hora Fichada',
        'origen_fichada': 'Origen Fichada',
    }

    def get_dni(self, obj):
        return obj.operario.dni
    get_dni.short_description = 'DNI'

    def get_nombre(self, obj):
        return obj.operario.nombre
    get_nombre.short_description = 'Nombre'

    def get_apellido(self, obj):
        return obj.operario.apellido
    get_apellido.short_description = 'Apellido'

    def formatted_hora_fichada(self, obj):
        return obj.hora_fichada.strftime('%d/%m/%Y %H:%M:%S')
    formatted_hora_fichada.short_description = 'Hora Fichada'

    def save_model(self, request, obj, form, change):
        if not change:  # Solo si es un nuevo registro
            obj.origen_fichada = 'Manual'
        obj.save()

    def mostrar_inconsistencia(self, obj):
        if obj.inconsistencia:
            return format_html('<span style="color: red; font-weight: bold;">Sí</span>')
        else:
            return format_html('<span style="color: green;">No</span>')
    mostrar_inconsistencia.short_description = 'Inconsistencia'
    mostrar_inconsistencia.admin_order_field = 'inconsistencia'

    def mostrar_valido(self, obj):
        if obj.valido:
            return format_html('<span style="color: green; font-weight: bold;">Sí</span>')
        else:
            return format_html('<span style="color: red;">No</span>')
    mostrar_valido.short_description = 'Válido'
    mostrar_valido.admin_order_field = 'valido'

    def view_history_button(self, obj):
        """Mostrar un botón para ver el historial del registro"""
        if obj.pk:
            url = reverse('admin:reloj_fichador_historicalregistrodiario_changelist') + f"?id={obj.pk}"
            return format_html(
                '<a class="button" href="{}">Ver Historial</a>',
                url
            )
        return ""
    view_history_button.short_description = "Historial"

    def generar_reporte(self, request, queryset):
        registros = list(queryset)  # Convertir a lista para poder trabajar con slicing
        total_registros = RegistroDiario.objects.count()  # Total sin filtrar
        items_por_pagina = 30  # Limitar a 30 ítems por página

        # Definir los encabezados manualmente
        encabezados = ['DNI', 'Nombre', 'Apellido','Hora Fichada', 'Tipo Movimiento', 'Origen Fichada']

        # Filas con valores de cada campo
        filas = []
        for registro in registros:
            fila = [
                registro.operario.dni,
                registro.operario.nombre,
                registro.operario.apellido,
                registro.hora_fichada.strftime('%d/%m/%Y %H:%M:%S'),
                registro.tipo_movimiento.replace('_', ' ').capitalize(),
                registro.origen_fichada.capitalize(),
            ]
            filas.append(fila)

        # Dividir las filas en páginas de 30 ítems
        page_filas = [filas[i:i + items_por_pagina] for i in range(0, len(filas), items_por_pagina)]
        total_pages = len(page_filas)

        # Definir el título dinámico
        titulo_reporte = "Reporte de Registro Diario"


        # Renderizar el template HTML con los datos dinámicos
        context = {
            'encabezados': encabezados,
            'page_filas': page_filas,  # Pasamos las filas ya divididas por páginas
            'num_columnas': len(encabezados),
            'current_date': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'current_page': 1,  # Este valor se puede manejar mejor dentro del template
            'total_pages': total_pages,  # Total de páginas
            'num_items_filtrados': len(registros),  # Número de registros filtrados
            'total_items': total_registros,
            'titulo_reporte': titulo_reporte          # Número total de registros
        }

        return render(request, 'reloj_fichador/reporte.html', context)


    def exportar_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse

        registros = list(queryset)  # Convertir a lista para trabajar fácilmente
        total_registros = RegistroDiario.objects.count()  # Total sin filtrar
        
        # Definir los encabezados manualmente
        encabezados = ['DNI', 'Nombre', 'Apellido', 'Hora Fichada', 'Tipo Movimiento', 'Origen Fichada']
        
        # Crear un nuevo archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Registro Diario"
        
        # Escribir los encabezados en la primera fila
        sheet.append(encabezados)
        
        # Filas con valores de cada campo
        for registro in registros:
            fila = [
                registro.operario.dni,
                registro.operario.nombre,
                registro.operario.apellido,
                registro.hora_fichada.strftime('%d/%m/%Y %H:%M:%S'),
                registro.tipo_movimiento.replace('_', ' ').capitalize(),
                registro.origen_fichada.capitalize(),
            ]
            sheet.append(fila)  # Añadir cada fila al archivo Excel
        
        # Preparar la respuesta HTTP para la descarga del archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_de_Registro_Diario.xlsx"'
        
        # Guardar el archivo Excel en la respuesta
        workbook.save(response)
        
        return response

    exportar_excel.short_description = "Exportar a Excel"

@admin.register(Horas_trabajadas)
class HorasTrabajadasAdmin(ExportMixin, admin.ModelAdmin):
    list_display = ('operario', 'fecha', 'get_horas_normales', 'get_horas_nocturnas')
    search_fields = ('operario__dni', 'operario__nombre', 'operario__apellido')
    list_filter = ('fecha', ('fecha', DateRangeFilter))
    actions = ['generar_reporte', 'exportar_excel', 'exportar_pdf']

    def get_queryset(self, request):
        """
        Filtra los registros para ocultar aquellos con 0 horas trabajadas y 0 horas nocturnas.
        Optimiza la consulta con select_related.
        """
        queryset = super().get_queryset(request).select_related('operario')
        return queryset.exclude(horas_normales=timedelta(0), horas_nocturnas=timedelta(0))

    def get_horas_normales(self, obj):
        total_seconds = obj.horas_normales.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        return f"{hours}h {minutes}m"

    get_horas_normales.short_description = 'Horas Normales'

    def get_horas_nocturnas(self, obj):
        total_seconds = obj.horas_nocturnas.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        return f"{hours}h {minutes}m"

    get_horas_nocturnas.short_description = 'Horas Nocturnas'

    def generar_reporte(self, request, queryset):
        registros = list(queryset)
        total_registros = Horas_trabajadas.objects.count()
        items_por_pagina = 30

        encabezados = ['DNI', 'Nombre', 'Apellido', 'Fecha', 'Horas Trabajadas', 'Horas Nocturnas']
        filas = []
        for registro in registros:
            fila = [
                registro.operario.dni,
                registro.operario.nombre,
                registro.operario.apellido,
                registro.fecha.strftime('%d/%m/%Y'),
                self.get_horas_normales(registro),
                self.get_horas_nocturnas(registro),
            ]
            filas.append(fila)

        # Dividir las filas en páginas de 30 ítems
        page_filas = [filas[i:i + items_por_pagina] for i in range(0, len(filas), items_por_pagina)]
        total_pages = len(page_filas)

        # Definir el título dinámico
        titulo_reporte = "Reporte de Horas Trabajadas"

        context = {
            'encabezados': encabezados,
            'page_filas': page_filas,
            'num_columnas': len(encabezados),
            'current_date': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'current_page': 1,
            'total_pages': total_pages,
            'num_items_filtrados': len(registros),
            'total_items': total_registros,
            'titulo_reporte': titulo_reporte
        }

        return render(request, 'reloj_fichador/reporte.html', context)

    def exportar_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse

        registros = list(queryset)
        encabezados = ['DNI', 'Nombre', 'Apellido', 'Fecha', 'Horas Trabajadas', 'Horas Nocturnas']
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Horas Trabajadas"
        sheet.append(encabezados)
        
        for registro in registros:
            fila = [
                registro.operario.dni,
                registro.operario.nombre,
                registro.operario.apellido,
                registro.fecha.strftime('%d/%m/%Y'),
                self.get_horas_normales(registro),
                self.get_horas_nocturnas(registro),
            ]
            sheet.append(fila)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Horas_Trabajadas.xlsx"'
        workbook.save(response)
        
        return response

    exportar_excel.short_description = "Exportar a Excel"

    def exportar_pdf(self, request, queryset):
        # Usar la función generar_pdf con cálculo automático de totales de horas
        return exportar_pdf(self, request, queryset, calculate_hours_total=True)


@admin.register(Horas_extras)
class HorasExtrasAdmin(ExportMixin, admin.ModelAdmin):
    list_display = ('operario', 'fecha', 'get_horas_extras')
    search_fields = ('operario__dni', 'operario__nombre', 'operario__apellido')
    list_filter = ('fecha', ('fecha', DateRangeFilter))
    actions = ['generar_reporte', 'exportar_excel', 'exportar_pdf']

    def get_queryset(self, request):
        """
        Filtra los registros para ocultar aquellos con 0 horas extras.
        Optimiza la consulta con select_related.
        """
        queryset = super().get_queryset(request).select_related('operario')
        return queryset.exclude(horas_extras=timedelta(0))


    def get_horas_extras(self, obj):
        total_seconds = obj.horas_extras.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        return f"{hours}h {minutes}m"

    get_horas_extras.short_description = 'Horas Extras'

    def generar_reporte(self, request, queryset):
        registros = list(queryset)
        total_registros = Horas_extras.objects.count()
        items_por_pagina = 30

        encabezados = ['DNI', 'Nombre', 'Apellido', 'Fecha', 'Horas Extras']
        filas = []
        for registro in registros:
            fila = [
                registro.operario.dni,
                registro.operario.nombre,
                registro.operario.apellido,
                registro.fecha.strftime('%d/%m/%Y'),
                self.get_horas_extras(registro),
            ]
            filas.append(fila)

        # Dividir las filas en páginas de 30 ítems
        page_filas = [filas[i:i + items_por_pagina] for i in range(0, len(filas), items_por_pagina)]
        total_pages = len(page_filas)

         # Definir el título dinámico
        titulo_reporte = "Reporte de Horas Extras"

        context = {
            'encabezados': encabezados,
            'page_filas': page_filas,
            'num_columnas': len(encabezados),
            'current_date': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'current_page': 1,
            'total_pages': total_pages,
            'num_items_filtrados': len(registros),
            'total_items': total_registros,
            'titulo_reporte': titulo_reporte
        }

        return render(request, 'reloj_fichador/reporte.html', context)

    def exportar_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse

        registros = list(queryset)
        encabezados = ['DNI', 'Nombre', 'Apellido', 'Fecha', 'Horas Extras']
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Horas Extras"
        sheet.append(encabezados)
        
        for registro in registros:
            fila = [
                registro.operario.dni,
                registro.operario.nombre,
                registro.operario.apellido,
                registro.fecha.strftime('%d/%m/%Y'),
                self.get_horas_extras(registro),
            ]
            sheet.append(fila)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Horas_Extras.xlsx"'
        workbook.save(response)
        
        return response

    exportar_excel.short_description = "Exportar a Excel"

    def exportar_pdf(self, request, queryset):
        # Usar la función generar_pdf con cálculo automático de totales de horas
        return exportar_pdf(self, request, queryset, calculate_hours_total=True)



@admin.register(Horas_totales)
class HorasTotalesAdmin(ExportMixin, admin.ModelAdmin):
    list_display = ('get_dni', 'operario', 'get_horas_normales', 'get_horas_nocturnas', 'get_horas_extras', 'get_horas_feriado')
    search_fields = ('operario__dni', 'operario__nombre', 'operario__apellido')
    list_filter = ('mes_actual',)
    actions = ['generar_reporte', 'exportar_excel', 'exportar_pdf']

    def get_queryset(self, request):
        """
        Filtra los registros para ocultar aquellos con 0 horas normales, nocturnas, extras y feriado.
        Optimiza la consulta con select_related.
        """
        queryset = super().get_queryset(request).select_related('operario')
        return queryset.exclude(
            horas_normales=timedelta(0),
            horas_nocturnas=timedelta(0),
            horas_extras=timedelta(0),
            horas_feriado=timedelta(0)
        )

    def get_dni(self, obj):
        return obj.operario.dni
    get_dni.short_description = 'DNI'

    def get_horas_normales(self, obj):
        total_seconds = obj.horas_normales.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        return f"{hours}h {minutes}m"
    get_horas_normales.short_description = 'Horas Normales'

    def get_horas_nocturnas(self, obj):
        total_seconds = obj.horas_nocturnas.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        return f"{hours}h {minutes}m"
    get_horas_nocturnas.short_description = 'Horas Nocturnas'

    def get_horas_extras(self, obj):
        total_seconds = obj.horas_extras.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        return f"{hours}h {minutes}m"
    get_horas_extras.short_description = 'Horas Extras'

    def get_horas_feriado(self, obj):
        total_seconds = obj.horas_feriado.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        return f"{hours}h {minutes}m"
    get_horas_feriado.short_description = 'Horas Feriado'

    def generar_reporte(self, request, queryset):
        registros = list(queryset)
        total_registros = Horas_totales.objects.count()
        items_por_pagina = 30

        encabezados = ['DNI', 'Nombre', 'Apellido', 'Horas Normales', 'Horas Nocturnas', 'Horas Extras', 'Horas Feriado']
        filas = []
        for registro in registros:
            fila = [
                registro.operario.dni,
                registro.operario.nombre,
                registro.operario.apellido,
                self.get_horas_normales(registro),
                self.get_horas_nocturnas(registro),
                self.get_horas_extras(registro),
                self.get_horas_feriado(registro),
            ]
            filas.append(fila)

        # Dividir las filas en páginas de 30 ítems
        page_filas = [filas[i:i + items_por_pagina] for i in range(0, len(filas), items_por_pagina)]
        total_pages = len(page_filas)

        # Definir el título dinámico
        titulo_reporte = "Reporte de Horas Extras"

        context = {
            'encabezados': encabezados,
            'page_filas': page_filas,
            'num_columnas': len(encabezados),
            'current_date': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'current_page': 1,
            'total_pages': total_pages,
            'num_items_filtrados': len(registros),
            'total_items': total_registros,
            'titulo_reporte': titulo_reporte
        }

        return render(request, 'reloj_fichador/reporte.html', context)

    def exportar_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse

        registros = list(queryset)
        encabezados = ['DNI', 'Nombre', 'Apellido', 'Horas Normales', 'Horas Nocturnas', 'Horas Extras', 'Horas Feriado']
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Horas Totales"
        sheet.append(encabezados)
        
        for registro in registros:
            fila = [
                registro.operario.dni,
                registro.operario.nombre,
                registro.operario.apellido,
                self.get_horas_normales(registro),
                self.get_horas_nocturnas(registro),
                self.get_horas_extras(registro),
                self.get_horas_feriado(registro),
            ]
            sheet.append(fila)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Horas_Totales.xlsx"'
        workbook.save(response)
        
        return response

    exportar_excel.short_description = "Exportar a Excel"

    def exportar_pdf(self, request, queryset):
        # Usar la función generar_pdf con cálculo automático de totales de horas
        return exportar_pdf(self, request, queryset, calculate_hours_total=True)


@admin.register(RegistroAsistencia)
class RegistroAsistenciaAdmin(ExportMixin, admin.ModelAdmin):
    list_display = (
        'operario', 'fecha', 'estado_asistencia', 'estado_justificacion_selector', 'descripcion', 'acciones'
    )
    list_filter = ('estado_asistencia', 'estado_justificacion', 'fecha')
    search_fields = ('operario__dni', 'operario__nombre', 'operario__apellido')
    actions = ['marcar_justificado', 'marcar_no_justificado', 'generar_reporte', 'exportar_excel', 'exportar_pdf']

    def estado_justificacion_selector(self, obj):
        return '✅' if obj.estado_justificacion else '❌'
    estado_justificacion_selector.short_description = 'Justificación'

    def marcar_justificado(self, request, queryset):
        queryset.update(estado_justificacion=True)
        self.message_user(request, "Las ausencias seleccionadas han sido marcadas como justificadas.")

    def marcar_no_justificado(self, request, queryset):
        queryset.update(estado_justificacion=False)
        self.message_user(request, "Las ausencias seleccionadas han sido marcadas como no justificadas.")

    def acciones(self, obj):
        return format_html(
            '<a class="button" href="{}">Cargar Licencia</a>',
            reverse('admin:cargar_licencia', args=[obj.pk])
        )
    acciones.short_description = 'Acciones'

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('cargar-licencia/<int:pk>/', self.cargar_licencia, name='cargar_licencia'),
        ]
        return custom_urls + urls

    def cargar_licencia(self, request, pk):
        registro_asistencia = get_object_or_404(RegistroAsistencia, pk=pk)

        if request.method == 'POST':
            form = LicenciaForm(request.POST, request.FILES)
            if form.is_valid():
                licencia = form.save(commit=False)
                licencia.operario = registro_asistencia.operario
                licencia.save()
                registro_asistencia.estado_justificacion = True
                registro_asistencia.descripcion = form.cleaned_data['descripcion']
                registro_asistencia.save()
                return redirect('admin:reloj_fichador_registroasistencia_changelist')
        else:
            form = LicenciaForm()

        return render(request, 'admin/cargar_licencia.html', {'form': form, 'registro_asistencia': registro_asistencia})

    def generar_reporte(self, request, queryset):
        registros = list(queryset)
        total_registros = RegistroAsistencia.objects.count()
        items_por_pagina = 20

        encabezados = ['Operario', 'Fecha', 'Estado Asistencia', 'Justificación', 'Descripción']

        filas = []
        for registro in registros:
            fila = [
                registro.operario,
                registro.fecha.strftime('%d/%m/%Y'),
                registro.estado_asistencia.capitalize(),
                '✅' if registro.estado_justificacion else '❌',
                registro.descripcion or ''
            ]
            filas.append(fila)

        # Dividir las filas en páginas de 30 ítems
        page_filas = [filas[i:i + items_por_pagina] for i in range(0, len(filas), items_por_pagina)]
        total_pages = len(page_filas)

        # Definir el título dinámico
        titulo_reporte = "Reporte de Asistencias"

        context = {
            'encabezados': encabezados,
            'page_filas': page_filas,
            'num_columnas': len(encabezados),
            'current_date': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'current_page': 1,
            'total_pages': total_pages,
            'num_items_filtrados': len(registros),
            'total_items': total_registros,
            'titulo_reporte': titulo_reporte
        }

        return render(request, 'reloj_fichador/reporte.html', context)

    def exportar_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse

        registros = list(queryset)
        encabezados = ['Operario', 'Fecha', 'Estado Asistencia', 'Justificación', 'Descripción']

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Registro de Asistencia"
        sheet.append(encabezados)

        for registro in registros:
            fila = [
                str(registro.operario),
                registro.fecha.strftime('%d/%m/%Y'),
                registro.estado_asistencia.capitalize(),
                '✅' if registro.estado_justificacion else '❌',
                registro.descripcion or ''
            ]
            sheet.append(fila)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_de_Registro_Asistencia.xlsx"'
        workbook.save(response)

        return response

    exportar_excel.short_description = "Exportar a Excel"

    def exportar_pdf(self, request, queryset):
        # Usar la función generar_pdf con cálculo automático de totales de horas
        return exportar_pdf(self, request, queryset, calculate_hours_total=True)

@admin.register(Horario)
class HorarioAdmin(ExportarPDFMixin, admin.ModelAdmin):
    list_display = ('nombre', 'hora_inicio', 'hora_fin')
    search_fields = ('nombre',)

@admin.register(Area)
class AreaAdmin(ExportarPDFMixin, admin.ModelAdmin):
    list_display = ('nombre',)
    search_fields = ('nombre',)
    filter_horizontal = ('horarios',)

@admin.register(Horas_feriado)
class HorasFeriadoAdmin(ExportMixin, admin.ModelAdmin):
    list_display = ('operario', 'fecha', 'horas_feriado')
    search_fields = ('operario__dni', 'operario__nombre', 'operario__apellido')
    list_filter = ('fecha', ('fecha', DateRangeFilter))
    actions = ['exportar_pdf', 'exportar_excel']

    def exportar_pdf(self, request, queryset):
        # Usar la función generar_pdf con cálculo automático de totales de horas
        return exportar_pdf(self, request, queryset, calculate_hours_total=True)

@admin.register(LogEntry)
class LogEntryAdmin(ExportarPDFMixin, admin.ModelAdmin):
    list_display = ('action_time', 'user', 'content_type', 'object_repr', 'action_flag', 'change_message')
    list_filter = ('action_flag', 'user', 'content_type')
    search_fields = ('object_repr', 'change_message', 'user__username')
    readonly_fields = ('action_time', 'user', 'content_type', 'object_repr', 'action_flag', 'change_message')
    
    def has_add_permission(self, request):
        return False  # Evita que se puedan añadir nuevos registros desde el admin
    
    def has_change_permission(self, request, obj=None):
        return False  # Evita cambios en los registros

    def has_delete_permission(self, request, obj=None):
        return False  # Evita eliminación de registros

    def action_flag(self, obj):
        """
        Muestra una representación más amigable de las acciones.
        """
        if obj.action_flag == 1:
            return format_html('<span style="color:green;">Creación</span>')
        elif obj.action_flag == 2:
            return format_html('<span style="color:orange;">Edición</span>')
        elif obj.action_flag == 3:
            return format_html('<span style="color:red;">Eliminación</span>')
        return obj.action_flag

    action_flag.short_description = 'Acción'

# Registro de modelos históricos
@admin.register(HistoricalOperario)
class HistoricalOperarioAdmin(admin.ModelAdmin):
    list_display = ('dni', 'nombre', 'apellido', 'activo', 'history_date', 'history_user', 'history_type')
    list_filter = ('history_date', 'history_type', 'activo')
    search_fields = ('dni', 'nombre', 'apellido', 'history_user__username')
    readonly_fields = ('dni', 'nombre', 'apellido', 'history_date', 'history_user', 'history_type')
    
    def has_add_permission(self, request):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False

@admin.register(HistoricalRegistroDiario)
class HistoricalRegistroDiarioAdmin(admin.ModelAdmin):
    list_display = ('get_operario', 'tipo_movimiento', 'formatted_hora_fichada', 'valido', 'inconsistencia', 'history_date', 'history_user', 'history_type')
    list_filter = ('history_date', 'history_type', 'tipo_movimiento', 'valido', 'inconsistencia')
    search_fields = ('operario__dni', 'operario__nombre', 'operario__apellido', 'history_user__username')
    readonly_fields = ('tipo_movimiento', 'hora_fichada', 'valido', 'inconsistencia', 'history_date', 'history_user', 'history_type')
    
    def formatted_hora_fichada(self, obj):
        return obj.hora_fichada.strftime('%d/%m/%Y %H:%M:%S') if obj.hora_fichada else '—'
    formatted_hora_fichada.short_description = 'Hora Fichada'
    
    def get_operario(self, obj):
        return f"{obj.operario}" if obj.operario else '—'
    get_operario.short_description = 'Operario'
    
    def has_add_permission(self, request):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False

# Asignar la función exportar_pdf a cada clase
OperarioAdmin.exportar_pdf = exportar_pdf
RegistroDiarioAdmin.exportar_pdf = exportar_pdf  # Usar la versión sin cálculo de totales
HorarioAdmin.exportar_pdf = exportar_pdf
AreaAdmin.exportar_pdf = exportar_pdf
LogEntryAdmin.exportar_pdf = exportar_pdf
RegistroAsistenciaAdmin.exportar_pdf = exportar_pdf

# Personalizar exportar_pdf para añadir totales en reportes de horas
def exportar_pdf_con_totales(modeladmin, request, queryset):
    """Función específica para exportar reportes con totales de horas"""
    return exportar_pdf(modeladmin, request, queryset, calculate_hours_total=True)

exportar_pdf_con_totales.short_description = "Exportar seleccionados a PDF (con totales)"

# Asignar la función personalizada SOLO a las clases de horas
HorasTrabajadasAdmin.exportar_pdf = exportar_pdf_con_totales
HorasExtrasAdmin.exportar_pdf = exportar_pdf_con_totales
HorasTotalesAdmin.exportar_pdf = exportar_pdf_con_totales
HorasFeriadoAdmin.exportar_pdf = exportar_pdf_con_totales

# Personalizar el admin de User para limitar permisos
class RestrictedUserAdmin(UserAdmin):
    """
    Administrador personalizado para User que restringe la edición
    de usuarios de manera que los usuarios staff solo puedan editar
    su propio perfil, mientras que los superusuarios pueden editar cualquiera.
    """
    
    def has_change_permission(self, request, obj=None):
        # Si el usuario es superusuario, tiene permiso completo
        if request.user.is_superuser:
            return True
        
        # Si estamos comprobando permisos generales (sin objeto específico)
        if obj is None:
            return True
        
        # Los usuarios solo pueden modificar sus propios datos
        return obj == request.user
    
    def get_list_filter(self, request):
        # Mostrar filtros solo a superusuarios
        if request.user.is_superuser:
            return super().get_list_filter(request)
        return []
    
    def get_queryset(self, request):
        # Superusuarios ven todos los usuarios, el resto solo se ve a sí mismo
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(pk=request.user.pk)
    
    def changelist_view(self, request, extra_context=None):
        # Si el usuario no es superusuario, redirigir directamente a su página de edición
        if not request.user.is_superuser:
            return HttpResponseRedirect(
                reverse('admin:auth_user_change', args=(request.user.id,))
            )
        return super().changelist_view(request, extra_context)
    
    def has_module_permission(self, request):
        # Todos los usuarios staff pueden ver el módulo
        return request.user.is_staff
    
    def has_add_permission(self, request):
        # Solo superusuarios pueden añadir usuarios
        return request.user.is_superuser
    
    def has_delete_permission(self, request, obj=None):
        # Solo superusuarios pueden eliminar usuarios
        return request.user.is_superuser
    
    def get_fieldsets(self, request, obj=None):
        # Definir los campos visibles según el tipo de usuario
        if not obj:
            return super().get_fieldsets(request, obj)
        
        # Si es superusuario, mostrar todos los campos
        if request.user.is_superuser:
            return super().get_fieldsets(request, obj)
        
        # Para usuarios normales, limitar los campos visibles
        # Eliminar campos de permisos para usuarios no superusuarios
        return [
            (None, {'fields': ('username', 'password')}),
            ('Información personal', {'fields': ('first_name', 'last_name', 'email')}),
        ]

# Desregistrar el UserAdmin predeterminado y registrar nuestro RestrictedUserAdmin
admin.site.unregister(User)
admin.site.register(User, RestrictedUserAdmin)